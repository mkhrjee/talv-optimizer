"""Build a TALVs.xlsx-style workbook from sweep results.

Layout per 4-part sheet (matches the analysts' current file):
  * A1:D  -> summary table: TALV, Lineholders, Reserves, Open Time (%)
  * G1:H  -> Employee#, PlannedAbsenceCredit
  * I1..  -> one column per TALV ("TALV: xx.x"): per-pilot sequence credit,
             replaced with 8888 where the pilot would fall to Reserve.
  * A line chart of Lineholders / Reserves / Open Time (%) over the TALV range.
"""

from __future__ import annotations

from typing import Iterable, List

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .optimizer import GroupResult

RESERVE_SENTINEL = 8888

# American Airlines palette.
AA_BLUE = "0078D2"
AA_RED = "C30019"
AA_DARK = "45596A"
HEADER_FILL = PatternFill("solid", fgColor=AA_BLUE)
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _sheet_name(result: GroupResult) -> str:
    fp = result.four_part
    # e.g. 777CALAXI -> EQP=777, SEAT=CA, BASE=LAX, DIV=I  -> "LAX 777 CA"
    if len(fp) >= 8:
        eqp, seat, base = fp[0:3], fp[3:5], fp[5:8]
        name = f"{base} {eqp} {seat}"
    else:
        name = fp
    return name[:31]  # Excel sheet-name limit


def _style_header(cell) -> None:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center")


def _write_group_sheet(wb: Workbook, result: GroupResult) -> None:
    ws = wb.create_sheet(title=_sheet_name(result))

    # --- Summary table (A:D) ---
    headers = ["TALV", "Lineholders", "Reserves", "Open Time (%)"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        _style_header(c)

    for i, talv in enumerate(result.talvs):
        row = i + 2
        ws.cell(row=row, column=1, value=talv)
        ws.cell(row=row, column=2, value=result.lineholders[i])
        ws.cell(row=row, column=3, value=result.reserves[i])
        ws.cell(row=row, column=4, value=result.open_times[i])

    n_talv = len(result.talvs)

    # --- Per-pilot tracker (G onward) ---
    base_col = 7  # column G
    ws.cell(row=1, column=base_col, value="Employee#")
    _style_header(ws.cell(row=1, column=base_col))
    ws.cell(row=1, column=base_col + 1, value="PlannedAbsenceCredit")
    _style_header(ws.cell(row=1, column=base_col + 1))

    talv_keys = [f"TALV: {t}" for t in result.talvs]
    for j, key in enumerate(talv_keys):
        _style_header(ws.cell(row=1, column=base_col + 2 + j, value=key))

    for i, emp in enumerate(result.employees):
        row = i + 2
        ws.cell(row=row, column=base_col, value=emp)
        ws.cell(row=row, column=base_col + 1, value=result.planned_absence[i])
        for j, key in enumerate(talv_keys):
            reserve = result.reserve_flag[key][i]
            value = RESERVE_SENTINEL if reserve else round(result.tracker[key][i], 2)
            ws.cell(row=row, column=base_col + 2 + j, value=value)

    # --- Chart ---
    chart = LineChart()
    chart.title = f"{result.four_part} — Lineholders / Reserves / Open Time over TALV"
    chart.style = 2
    chart.height = 10
    chart.width = 24
    chart.y_axis.title = "Number of Pilots"
    chart.x_axis.title = "TALV"

    cats = Reference(ws, min_col=1, min_row=2, max_row=n_talv + 1)
    for col in (2, 3):  # Lineholders, Reserves
        data = Reference(ws, min_col=col, min_row=1, max_row=n_talv + 1)
        chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # Open Time (%) on a secondary axis.
    open_chart = LineChart()
    open_data = Reference(ws, min_col=4, min_row=1, max_row=n_talv + 1)
    open_chart.add_data(open_data, titles_from_data=True)
    open_chart.set_categories(cats)
    open_chart.y_axis.axId = 200
    open_chart.y_axis.title = "Open Time (%)"
    open_chart.y_axis.crosses = "max"
    chart += open_chart

    anchor_col = get_column_letter(base_col + 2 + len(talv_keys) + 1)
    ws.add_chart(chart, f"{anchor_col}2")

    # Freeze header + widen key columns.
    ws.freeze_panes = "A2"
    ws.column_dimensions[get_column_letter(base_col + 1)].width = 20


def build_workbook(results: Iterable[GroupResult], out_path: str) -> str:
    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet
    for result in results:
        _write_group_sheet(wb, result)
    if not wb.sheetnames:
        wb.create_sheet(title="No Data")
    wb.save(out_path)
    return out_path
